import requests
from bs4 import BeautifulSoup
import time
import os
from openpyxl import Workbook, load_workbook

def get_page_data(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    property_items = soup.find_all('div', class_='_475e888a')
    data = []

    for property_item in property_items:
        try:
            price_elem = property_item.find('div', class_='_2923a568')
            price = price_elem.text.strip().replace(',', '').replace('EGP', '').strip() if price_elem else 'N/A'
            price = float(price) if price != 'N/A' else 0
        except Exception as e:
            price = 0
            print(f"Error extracting price: {e}")

        try:
            property_type_elem = property_item.find('span', class_='_19e94678 e0abc2de')
            property_type = property_type_elem.text.strip() if property_type_elem else 'N/A'
        except Exception as e:
            property_type = 'Error extracting property type'
            print(f"Error extracting property type: {e}")

        try:
            location_elem = property_item.find('h3', class_='_4402bd70')
            location = location_elem.text.strip() if location_elem else 'N/A'
            location_parts = location.split(', ')
            neighborhood = location_parts[0] if len(location_parts) > 0 else 'N/A'
            district = location_parts[1] if len(location_parts) > 1 else 'N/A'
            city = location_parts[2] if len(location_parts) > 2 else 'N/A'
        except Exception as e:
            neighborhood = district = city = 'Error extracting location'
            print(f"Error extracting location: {e}")

        try:
            no_of_rooms_elem = property_item.find('span', {'aria-label': 'Beds'})
            no_of_rooms = no_of_rooms_elem.text.strip() if no_of_rooms_elem else 'N/A'
        except Exception as e:
            no_of_rooms = 'Error extracting number of rooms'
            print(f"Error extracting number of rooms: {e}")

        try:
            no_of_toilets_elem = property_item.find('span', {'aria-label': 'Baths'})
            no_of_toilets = no_of_toilets_elem.text.strip() if no_of_toilets_elem else 'N/A'
        except Exception as e:
            no_of_toilets = 'Error extracting number of toilets'
            print(f"Error extracting number of toilets: {e}")

        try:
            area_elem = property_item.find('span', {'aria-label': 'Area'})
            area_text = area_elem.text.strip().replace(' Sq. M.', '').replace(',', '') if area_elem else 'N/A'
            area = float(area_text) if area_text != 'N/A' else 0
        except Exception as e:
            area = 0
            print(f"Error extracting area: {e}")

        try:
            desc_data = property_item.find('h2', {'aria-label': 'Title'})
            description = desc_data.text.strip() if desc_data else 'N/A'
        except Exception as e:
            description = 'Error extracting description'
            print(f"Error extracting description: {e}")

        avg_price_per_meter = price / area if area > 0 else 0

        data.append({
            'Price': price,
            'Type': property_type,
            'Neighborhood': neighborhood,
            'District': district,
            'City': city,
            'Number of Rooms': no_of_rooms,
            'Number of Toilets': no_of_toilets,
            'Area': area,
            'Average Price per Meter': avg_price_per_meter,
            'Description': description
        })
    return data

def scrape_all_pages(base_url, max_pages=150):
    all_data = []
    page_number = 1

    while page_number <= max_pages:
        print(f"Scraping page {page_number}...")
        if page_number == 1:
            url = base_url
        else:
            url = f"{base_url}page-{page_number}/"

        page_data = get_page_data(url)
        
        if not page_data:
            print("No more data found or end of pages reached.")
            break

        all_data.extend(page_data)
        page_number += 1
        time.sleep(2)  # Be respectful and avoid making too many requests too quickly

    return all_data

def update_excel_with_new_data(file_path, new_data):
    file_exists = os.path.isfile(file_path)

    if file_exists:
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        headers = ['Price', 'Type', 'Neighborhood', 'District', 'City', 'Number of Rooms', 'Number of Toilets', 'Area', 'Average Price per Meter', 'Average Price per Meter (District)', 'Description']
        ws.append(headers)

    existing_properties = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 11:
            existing_properties.add((row[1], row[2], row[3], row[4], row[10]))  # Type, Neighborhood, District, City, Description

    district_prices = {}
    district_areas = {}
    for row in new_data:
        district = row['District']
        if district not in district_prices:
            district_prices[district] = 0
            district_areas[district] = 0
        district_prices[district] += row['Price']
        district_areas[district] += row['Area']

    for row in new_data:
        district = row['District']
        avg_price_per_meter_district = district_prices[district] / district_areas[district] if district_areas[district] > 0 else 0
        key = (row['Type'], row['Neighborhood'], row['District'], row['City'], row['Description'])
        if key not in existing_properties:
            ws.append([row['Price'], row['Type'], row['Neighborhood'], row['District'], row['City'], row['Number of Rooms'], row['Number of Toilets'], row['Area'], row['Average Price per Meter'], avg_price_per_meter_district, row['Description']])

    wb.save(file_path)

# List of base URLs
base_urls = [
    # 'https://www.bayut.eg/en/cairo/properties-for-rent-in-new-cairo/',
    # 'https://www.bayut.eg/en/cairo/properties-for-rent-in-zamalek/',
    # 'https://www.bayut.eg/en/cairo/properties-for-rent-in-maadi/',
    # 'https://www.bayut.eg/en/cairo/properties-for-rent-in-madinaty/',
    # 'https://www.bayut.eg/en/cairo/properties-for-rent-in-mostakbal-city/',
    # 'https://www.bayut.eg/en/cairo/properties-for-rent-in-new-capital-city/',
    # 'https://www.bayut.eg/en/cairo/properties-for-rent-in-shorouk-city/',
    # 'https://www.bayut.eg/en/cairo/properties-for-rent-in-nasr-city/',
    # 'https://www.bayut.eg/en/cairo/properties-for-rent-in-heliopolis/',
    # 'https://www.bayut.eg/en/cairo/properties-for-rent-in-new-heliopolis/',
    # 'https://www.bayut.eg/en/cairo/properties-for-rent-in-katameya/',
    # 'https://www.bayut.eg/en/cairo/properties-for-rent-in-mokattam/',
    # 'https://www.bayut.eg/en/cairo/properties-for-rent-in-sheraton/',
    # 'https://www.bayut.eg/en/cairo/properties-for-rent-in-al-manial/',
    # 'https://www.bayut.eg/en/cairo/properties-for-rent-in-garden-city/'
    
    
    
]

# Initialize a list to hold all new data
all_new_data = []

# Scrape data from each URL
for base_url in base_urls:  
    new_data = scrape_all_pages(base_url, max_pages=1) # Change value of pages 
    all_new_data.extend(new_data)

# Update Excel with all new data
update_excel_with_new_data('rent_bayut.xlsx', all_new_data)

print("Scraping complete. Data updated in rent_bayut.xlsx.")
